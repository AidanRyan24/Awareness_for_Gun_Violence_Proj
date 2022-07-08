import openpyxl
from collections import Counter

path = 'shootingdata.xlsm'

wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 
  
# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column
  
print("Total Shootings this year:", row - 1)
print("Total Columns:", column)
  
# printing the value of first column
# Loop will print all values 
# of first column  
state = input('Select State: ')
data = {}
count = 0
#print("\nValue of first column")

for i in range(row - 1): 
    cell_obj = sheet_obj.cell(row = i + 1, column = 3) 
    if cell_obj.value == state:
        count += 1
    else:
        continue
    # if data.get(cell_obj.value) == True:
    #    data[cell_obj.value] += 1
    # else:
    #    data[cell_obj.value] = 1

print("There have been " + str(count) + " shootings in " + state)
#print(data)

      
# printing the value of first column
# Loop will print all values 
# of first row
# print("\nValue of first row")
# for i in range(1, column + 1): 
#     cell_obj = sheet_obj.cell(row = 2, column = i) 
#     print(cell_obj.value, end = " ")